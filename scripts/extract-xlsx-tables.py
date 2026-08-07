#!/usr/bin/env python3
"""Convert Excel worksheets into merged-cell-aware structured RAG tables."""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


@dataclass(frozen=True)
class WorkbookSpec:
    path: Path
    title: str
    revision: str


def parse_workbook(value: str) -> WorkbookSpec:
    parts = value.split("::")
    if len(parts) != 3:
        raise argparse.ArgumentTypeError("--workbook must be PATH::TITLE::REVISION")
    path = Path(parts[0]).expanduser().resolve()
    if not path.is_file():
        raise argparse.ArgumentTypeError(f"Excel file not found: {path}")
    return WorkbookSpec(path, parts[1].strip(), parts[2].strip())


def value_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).replace("\x00", "")).strip()


def markdown_table(matrix: list[list[str]]) -> str:
    if not matrix:
        return ""
    escaped = lambda value: (value or "원문상 빈칸").replace("|", "\\|").replace("\n", "<br>")
    return "\n".join([
        "| " + " | ".join(map(escaped, matrix[0])) + " |",
        "| " + " | ".join("---" for _ in matrix[0]) + " |",
        *("| " + " | ".join(map(escaped, row)) + " |" for row in matrix[1:]),
    ])


def worksheet_table(spec: WorkbookSpec, ws: Any, sheet_index: int) -> dict[str, Any] | None:
    populated = [cell for row in ws.iter_rows() for cell in row if cell.value not in (None, "")]
    if not populated:
        return None
    min_row = min(cell.row for cell in populated)
    max_row = max(cell.row for cell in populated)
    min_column = min(cell.column for cell in populated)
    max_column = max(cell.column for cell in populated)
    row_count = max_row - min_row + 1
    column_count = max_column - min_column + 1
    anchors = [["" for _ in range(column_count)] for _ in range(row_count)]
    expanded = [["" for _ in range(column_count)] for _ in range(row_count)]
    cells: list[dict[str, Any]] = []

    merged_by_anchor = {(area.min_row, area.min_col): area for area in ws.merged_cells.ranges}
    covered: set[tuple[int, int]] = set()
    for area in ws.merged_cells.ranges:
        for row in range(area.min_row, area.max_row + 1):
            for column in range(area.min_col, area.max_col + 1):
                covered.add((row, column))

    for row in range(min_row, max_row + 1):
        for column in range(min_column, max_column + 1):
            cell = ws.cell(row, column)
            if isinstance(cell, MergedCell):
                continue
            text = value_text(cell.value)
            area = merged_by_anchor.get((row, column))
            rowspan = area.max_row - area.min_row + 1 if area else 1
            colspan = area.max_col - area.min_col + 1 if area else 1
            if text or area:
                cells.append({
                    "row": row - min_row,
                    "column": column - min_column,
                    "rowspan": rowspan,
                    "colspan": colspan,
                    "text": text,
                    "coordinate": cell.coordinate,
                    "formula": text if text.startswith("=") else "",
                })
            anchors[row - min_row][column - min_column] = text
            for target_row in range(row, min(max_row + 1, row + rowspan)):
                for target_column in range(column, min(max_column + 1, column + colspan)):
                    expanded[target_row - min_row][target_column - min_column] = text

    title = f"[Excel 시트 {sheet_index}] {ws.title}"
    identity = f"{spec.path.name}\0{sheet_index}\0{json.dumps(expanded, ensure_ascii=False)}"
    checksum = hashlib.sha256(identity.encode("utf-8")).hexdigest()
    table_type = "form" if "양식" in spec.path.name else "data_table"
    search_text = "\n".join([
        spec.title,
        spec.path.name,
        ws.title,
        *(value for row in expanded for value in row if value),
    ])
    return {
        "id": f"KHT-{checksum[:24]}",
        "document_title": spec.title,
        "source_file": spec.path.name,
        "revision_basis": spec.revision,
        "page_start": sheet_index,
        "page_end": sheet_index,
        "table_index": 1,
        "table_title": title,
        "table_type": table_type,
        "context_before": f"Excel 원본 파일의 '{ws.title}' 시트",
        "context_after": f"실제 사용 영역 {ws.cell(min_row, min_column).coordinate}:{ws.cell(max_row, max_column).coordinate}",
        "row_count": row_count,
        "column_count": column_count,
        "cells": cells,
        "anchor_matrix": anchors,
        "expanded_matrix": expanded,
        "markdown": markdown_table(expanded),
        "search_text": search_text,
        "extraction_method": "text",
        "confidence": 1.0,
        "checksum_sha256": checksum,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", action="append", type=parse_workbook, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--summary", type=Path, required=True)
    args = parser.parse_args()
    tables: list[dict[str, Any]] = []
    documents: list[dict[str, Any]] = []
    for spec in args.workbook:
        workbook = load_workbook(spec.path, data_only=False, read_only=False)
        before = len(tables)
        for index, worksheet in enumerate(workbook.worksheets, 1):
            table = worksheet_table(spec, worksheet, index)
            if table:
                tables.append(table)
        documents.append({
            "document_title": spec.title,
            "source_file": spec.path.name,
            "revision_basis": spec.revision,
            "sheet_count": len(workbook.sheetnames),
            "table_count": len(tables) - before,
        })
    summary = {
        "schema_version": 1,
        "document_count": len(documents),
        "sheet_count": sum(item["sheet_count"] for item in documents),
        "table_count": len(tables),
        "cell_count": sum(len(table["cells"]) for table in tables),
        "documents": documents,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(tables, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    args.summary.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
